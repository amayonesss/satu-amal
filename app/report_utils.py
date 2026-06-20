import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import Program, User, Presensi
from app import db
from sqlalchemy import func

def generate_program_excel(bulan=None, tahun=None):
    wb = Workbook()
    
    # Setup styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="1F59D2")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    bold_data_font = Font(name=font_family, size=10, bold=True, color="333333")
    kpi_lbl_font = Font(name=font_family, size=9, color="555555", bold=True)
    kpi_val_font = Font(name=font_family, size=16, bold=True, color="1F59D2")
    
    header_fill = PatternFill(start_color="1F59D2", end_color="1F59D2", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFF", end_color="F8FAFF", fill_type="solid")
    kpi_fill = PatternFill(start_color="EAF0FD", end_color="EAF0FD", fill_type="solid")
    title_fill = PatternFill(start_color="0F3D99", end_color="0F3D99", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom = Border(bottom=Side(style='double', color='333333'), top=thin_border_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ----------------------------------------------------
    # SHEET 1: RINGKASAN EKSEKUTIF
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Ringkasan Eksekutif"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws1.merge_cells("A1:H2")
    title_cell = ws1["A1"]
    title_text = "LAPORAN REKAPITULASI PROGRAM — SATU AMAL"
    if bulan or tahun:
        period_text = f" ({bulan} {tahun})" if (bulan and tahun) else (f" ({bulan})" if bulan else f" (Tahun {tahun})")
        title_text = f"LAPORAN REKAPITULASI PROGRAM{period_text} — SATU AMAL"
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center
    
    # Fetch KPI Metrics
    prog_query = Program.query
    if bulan:
        prog_query = prog_query.filter(Program.bulan == bulan)
    if tahun:
        prog_query = prog_query.filter(Program.tahun == tahun)
    total_program = prog_query.count()
    
    pm_query = db.session.query(func.sum(Program.jumlah_penerima_manfaat))
    if bulan:
        pm_query = pm_query.filter(Program.bulan == bulan)
    if tahun:
        pm_query = pm_query.filter(Program.tahun == tahun)
    total_pm = pm_query.scalar() or 0
    
    lokasi_query = db.session.query(func.count(func.distinct(Program.kota))).filter(Program.kota != None, Program.kota != '')
    if bulan:
        lokasi_query = lokasi_query.filter(Program.bulan == bulan)
    if tahun:
        lokasi_query = lokasi_query.filter(Program.tahun == tahun)
    total_lokasi = lokasi_query.scalar() or 0
    
    total_relawan = User.query.filter_by(role='relawan', aktif=True).count()
    
    # Setup KPI Card Grid manually to avoid overlapping merge bounds
    # Card 1 (A4:B5) - Total Program
    ws1.merge_cells("A4:B4")
    ws1.merge_cells("A5:B5")
    ws1["A4"] = "TOTAL PROGRAM"
    ws1["A4"].font = kpi_lbl_font
    ws1["A4"].alignment = align_center
    ws1["A5"] = total_program
    ws1["A5"].font = kpi_val_font
    ws1["A5"].alignment = align_center
    
    # Card 2 (C4:D5) - Total PM
    ws1.merge_cells("C4:D4")
    ws1.merge_cells("C5:D5")
    ws1["C4"] = "TOTAL PENERIMA MANFAAT"
    ws1["C4"].font = kpi_lbl_font
    ws1["C4"].alignment = align_center
    ws1["C5"] = total_pm
    ws1["C5"].font = kpi_val_font
    ws1["C5"].alignment = align_center
    ws1["C5"].number_format = '#,##0'
    
    # Card 3 (E4:F5) - Total Lokasi
    ws1.merge_cells("E4:F4")
    ws1.merge_cells("E5:F5")
    ws1["E4"] = "KOTA DIJANGKAU"
    ws1["E4"].font = kpi_lbl_font
    ws1["E4"].alignment = align_center
    ws1["E5"] = total_lokasi
    ws1["E5"].font = kpi_val_font
    ws1["E5"].alignment = align_center
    
    # Card 4 (G4:H5) - Total Relawan
    ws1.merge_cells("G4:H4")
    ws1.merge_cells("G5:H5")
    ws1["G4"] = "RELAWAN AKTIF"
    ws1["G4"].font = kpi_lbl_font
    ws1["G4"].alignment = align_center
    ws1["G5"] = total_relawan
    ws1["G5"].font = kpi_val_font
    ws1["G5"].alignment = align_center
    
    # Apply backgrounds & borders to KPI cards
    for r in range(4, 6):
        for c in range(1, 9):
            cell = ws1.cell(row=r, column=c)
            cell.fill = kpi_fill
            cell.border = thin_border
            
    # Table 1: Kategori Breakdown (left side A8:D...)
    ws1["A7"] = "Analisis Kategori Program"
    ws1["A7"].font = section_font
    
    kat_headers = ["Kategori", "Program", "Total PM", "Persentase"]
    for idx, h in enumerate(kat_headers, start=1):
        cell = ws1.cell(row=8, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    kat_query = db.session.query(
        Program.kategori,
        func.count(Program.id).label('jml'),
        func.sum(Program.jumlah_penerima_manfaat).label('pm')
    ).filter(Program.kategori != None, Program.kategori != '')
    if bulan:
        kat_query = kat_query.filter(Program.bulan == bulan)
    if tahun:
        kat_query = kat_query.filter(Program.tahun == tahun)
    kat_rows = kat_query.group_by(Program.kategori).order_by(func.count(Program.id).desc()).all()
    
    current_row = 9
    total_prog_sum = sum(r[1] for r in kat_rows) or 1
    
    for row in kat_rows:
        ws1.cell(row=current_row, column=1, value=row[0] or 'Lainnya').font = data_font
        ws1.cell(row=current_row, column=2, value=row[1]).font = data_font
        ws1.cell(row=current_row, column=3, value=row[2] or 0).font = data_font
        pct = (row[1] / total_prog_sum)
        cell_pct = ws1.cell(row=current_row, column=4, value=pct)
        cell_pct.font = data_font
        cell_pct.number_format = '0.0%'
        
        for col_idx in range(1, 5):
            c = ws1.cell(row=current_row, column=col_idx)
            c.border = thin_border
            if current_row % 2 == 0:
                c.fill = zebra_fill
            if col_idx in [2, 3, 4]:
                c.alignment = align_right
                if col_idx == 3:
                    c.number_format = '#,##0'
            else:
                c.alignment = align_left
        current_row += 1
        
    # Total row for Kategori
    ws1.cell(row=current_row, column=1, value="Total").font = bold_data_font
    ws1.cell(row=current_row, column=1).alignment = align_left
    ws1.cell(row=current_row, column=1).border = double_bottom
    
    cell_prog = ws1.cell(row=current_row, column=2, value=total_prog_sum)
    cell_prog.font = bold_data_font
    cell_prog.alignment = align_right
    cell_prog.border = double_bottom
    
    cell_pm_sum = ws1.cell(row=current_row, column=3, value=total_pm)
    cell_pm_sum.font = bold_data_font
    cell_pm_sum.alignment = align_right
    cell_pm_sum.number_format = '#,##0'
    cell_pm_sum.border = double_bottom
    
    cell_pct_sum = ws1.cell(row=current_row, column=4, value=1.0)
    cell_pct_sum.font = bold_data_font
    cell_pct_sum.alignment = align_right
    cell_pct_sum.number_format = '0.0%'
    cell_pct_sum.border = double_bottom
    
    # Table 2: Sebaran Wilayah / Lokasi (right side F8:H...)
    ws1["F7"] = "Top 10 Sebaran Wilayah / Lokasi"
    ws1["F7"].font = section_font
    
    loc_headers = ["Kota / Wilayah", "Program", "Total PM"]
    for idx, h in enumerate(loc_headers, start=6):
        cell = ws1.cell(row=8, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    loc_query = db.session.query(
        Program.kota,
        func.count(Program.id).label('jml'),
        func.sum(Program.jumlah_penerima_manfaat).label('pm')
    ).filter(Program.kota != None, Program.kota != '')
    if bulan:
        loc_query = loc_query.filter(Program.bulan == bulan)
    if tahun:
        loc_query = loc_query.filter(Program.tahun == tahun)
    loc_rows = loc_query.group_by(Program.kota).order_by(func.sum(Program.jumlah_penerima_manfaat).desc()).limit(10).all()
    
    loc_row_idx = 9
    for row in loc_rows:
        ws1.cell(row=loc_row_idx, column=6, value=row[0]).font = data_font
        ws1.cell(row=loc_row_idx, column=7, value=row[1]).font = data_font
        ws1.cell(row=loc_row_idx, column=8, value=row[2] or 0).font = data_font
        
        for col_idx in range(6, 9):
            c = ws1.cell(row=loc_row_idx, column=col_idx)
            c.border = thin_border
            if loc_row_idx % 2 == 0:
                c.fill = zebra_fill
            if col_idx in [7, 8]:
                c.alignment = align_right
                if col_idx == 8:
                    c.number_format = '#,##0'
            else:
                c.alignment = align_left
        loc_row_idx += 1
        
    # Total row for Lokasi
    total_loc_prog = sum(r[1] for r in loc_rows)
    total_loc_pm = sum(r[2] for r in loc_rows)
    
    ws1.cell(row=loc_row_idx, column=6, value="Total Top 10").font = bold_data_font
    ws1.cell(row=loc_row_idx, column=6).alignment = align_left
    ws1.cell(row=loc_row_idx, column=6).border = double_bottom
    
    cell_loc_prog = ws1.cell(row=loc_row_idx, column=7, value=total_loc_prog)
    cell_loc_prog.font = bold_data_font
    cell_loc_prog.alignment = align_right
    cell_loc_prog.border = double_bottom
    
    cell_loc_pm = ws1.cell(row=loc_row_idx, column=8, value=total_loc_pm)
    cell_loc_pm.font = bold_data_font
    cell_loc_pm.alignment = align_right
    cell_loc_pm.number_format = '#,##0'
    cell_loc_pm.border = double_bottom
    
    # Adjust column widths for Sheet 1
    for col in ws1.columns:
        vals = []
        for cell in col:
            if cell.row > 3 and cell.value:
                vals.append(len(str(cell.value)))
        max_len = max(vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # ----------------------------------------------------
    # SHEET 2: DETAIL PROGRAM
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Detail Program")
    ws2.views.sheetView[0].showGridLines = True
    
    detail_headers = ["No", "Nama Program", "Kategori", "Lokasi", "Bulan", "Tahun", "Jumlah Relawan", "Target PM", "PM Aktual", "Catatan Evaluasi", "Kendala Lapangan"]
    for idx, h in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    prog_detail_query = Program.query
    if bulan:
        prog_detail_query = prog_detail_query.filter(Program.bulan == bulan)
    if tahun:
        prog_detail_query = prog_detail_query.filter(Program.tahun == tahun)
    programs = prog_detail_query.order_by(Program.id.desc()).all()
    
    row_idx = 2
    for idx, p in enumerate(programs, start=1):
        evaluasi = p.evaluasi_list[0] if p.evaluasi_list else None
        
        ws2.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws2.cell(row=row_idx, column=2, value=p.nama_program).alignment = align_left
        ws2.cell(row=row_idx, column=3, value=p.kategori or '-').alignment = align_center
        ws2.cell(row=row_idx, column=4, value=p.kota or '-').alignment = align_left
        ws2.cell(row=row_idx, column=5, value=p.bulan or '-').alignment = align_center
        ws2.cell(row=row_idx, column=6, value=p.tahun or '-').alignment = align_center
        ws2.cell(row=row_idx, column=7, value=p.jumlah_relawan or 0).alignment = align_right
        ws2.cell(row=row_idx, column=8, value=p.jumlah_penerima_manfaat or 0).alignment = align_right
        
        # Kolom evaluasi
        pm_aktual = evaluasi.jumlah_pm_aktual if evaluasi else 0
        catatan = evaluasi.catatan_keberhasilan if evaluasi else '-'
        kendala = evaluasi.kendala_lapangan if evaluasi else '-'
        
        ws2.cell(row=row_idx, column=9, value=pm_aktual).alignment = align_right
        ws2.cell(row=row_idx, column=10, value=catatan).alignment = align_left
        ws2.cell(row=row_idx, column=11, value=kendala).alignment = align_left
        
        # Apply styles
        for col_idx in range(1, 12):
            c = ws2.cell(row=row_idx, column=col_idx)
            c.font = data_font
            c.border = thin_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
            if col_idx in [7, 8, 9]:
                c.number_format = '#,##0'
        row_idx += 1
        
    # Add Total Row
    total_relawan_all = sum(p.jumlah_relawan or 0 for p in programs)
    total_pm_all = sum(p.jumlah_penerima_manfaat or 0 for p in programs)
    total_pm_aktual_all = sum((p.evaluasi_list[0].jumlah_pm_aktual if p.evaluasi_list else 0) for p in programs)
    
    ws2.cell(row=row_idx, column=1, value="TOTAL").font = bold_data_font
    ws2.cell(row=row_idx, column=1).alignment = align_center
    ws2.cell(row=row_idx, column=1).border = double_bottom
    
    for col_idx in range(2, 7):
        ws2.cell(row=row_idx, column=col_idx, value="").border = double_bottom
        
    cell_tot_rel = ws2.cell(row=row_idx, column=7, value=total_relawan_all)
    cell_tot_rel.font = bold_data_font
    cell_tot_rel.alignment = align_right
    cell_tot_rel.number_format = '#,##0'
    cell_tot_rel.border = double_bottom
    
    cell_tot_pm = ws2.cell(row=row_idx, column=8, value=total_pm_all)
    cell_tot_pm.font = bold_data_font
    cell_tot_pm.alignment = align_right
    cell_tot_pm.number_format = '#,##0'
    cell_tot_pm.border = double_bottom
    
    cell_tot_pm_aktual = ws2.cell(row=row_idx, column=9, value=total_pm_aktual_all)
    cell_tot_pm_aktual.font = bold_data_font
    cell_tot_pm_aktual.alignment = align_right
    cell_tot_pm_aktual.number_format = '#,##0'
    cell_tot_pm_aktual.border = double_bottom
    
    for col_idx in range(10, 12):
        ws2.cell(row=row_idx, column=col_idx, value="").border = double_bottom
    
    # Adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'B':
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 30)
        else:
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 11)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
